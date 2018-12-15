import openpyxl
import xlsxwriter

# this class is good if you have many columns
# and need to fill just a few of them with values
class ExcelFactory:

    def create_workbook(self, file):
        self.file = file
        self.workbook = openpyxl.Workbook()
        return self

    def add_sheet(self, sheet):
        self.worksheet = self.workbook.create_sheet(sheet)
        return self

    def add_header(self, header_list):
        self.header_list = header_list
        col_count = 1
        for header in header_list:
            self.worksheet.cell(row=1, column=col_count).value = header
            col_count += 1
        return self

    def add_values_to_column(self, header, value_list):
        for colx in range(1, self.worksheet.max_column+1):
            if header == self.worksheet.cell(row=1, column=colx).value:
                rowx = 1
                for value in value_list:
                    rowx += 1
                    self.worksheet.cell(row=rowx, column=colx).value = value
        return self

    def build(self):
        self.workbook.save(self.file)

# this class is good if you have many columns
# and need to fill all of them with values
class ExcelFactory2:

    def create_workbook(self, file):
        self.workbook = xlsxwriter.Workbook(file)

    def add_sheet(self, sheet_name):
        self.sheet = self.workbook.add_worksheet(sheet_name)

    def add_rows(self, rows):
        self._write_csv_to_sheet(self.sheet, rows)

    def _write_csv_to_sheet(self, sheet, csv):
        for i in range(0, len(csv)):
            sheet.write_row(i, 0, csv[i])
        self.workbook.close()

